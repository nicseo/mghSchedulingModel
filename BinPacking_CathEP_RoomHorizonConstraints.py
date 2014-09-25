"""
Created on Thursday, September 18, 2014
Modified on 9/25

@author: nicseo

This script models efficient scheduling of prime time procedures in both Cath and EP based
on data taken over a 125 day (25 week) period. The constraints of the model are:
    1. Room location (where the procedure must be done)
    2. Scheduling horizon (the time period during which the procedure must be scheduled)

Notes about use:
- You must have openpyxl package installed to run this script.

- The first part of this script sets up the necessary functions to carry out the analysis.
    At the bottom of this file, in the __main__ function, there are defined some variables
    that the user should verify before running the script (e.g. regarding the csv that the
    data is being read from).

- Results are output to an Excel sheet, whose workbook name and sheet must EXIST and
    be specified before running the script. Make sure the sheet is in the same directory
    as this script.

- Overflow procedures that are emergencies/same day procedures are listed during the day they
    were supposed to be scheduled. Overflows that were supposed to be scheduled within a week
    are listed in the first day of the week's overflows (Monday).
    
- If you change the number of Cath/EP rooms in the scheduling parameters, it is recommended
    that you CLEAR the output file of contents AND formatting before running.

"""
import csv
import os
import openpyxl as px
import copy


######################################################################################################
######################################################################################################
######################################### TIME PERIOD DATA TYPE ######################################
######################################################################################################
###################################################################################################### 


class TimePeriod:
    '''
    Class to model a given time period of scheduling.
    
    Initialization:
        TimePeriod(days,numCathRooms,numEPRooms)
            days - the number of days in the time period
            numCathRooms - the number of Cath rooms to be modeled
            numEPRooms - the number of EP rooms to be modeled

    User methods:
        packBins(procedures)

    Helper methods (not to be explicitly called outside of class):
        sumProcTimes(dataList)
        getTotalMinutesInLabForDay(day,lab)
        tryPlaceProcInLabDay(procedure,lab,day,nextOpenRoom,openRooms)
        packBinsForDay(day,daysProcedures)
        getTotalMinutesInLabForWeek(week,lab)
        tryPlaceProcInLabWeek(procedure,lab,week,nextOpenRoom,openRooms)
        packBinsForWeek(week,weeksProcedures)
        
    '''
    

    def __init__(self,days,numCathRooms,numEPRooms):

        CathRooms = {i:[] for i in xrange(numCathRooms)}
        EPRooms = {i:[] for i in xrange(numEPRooms)}
        procsNotPlaced = []

        self.dayBins = [[copy.deepcopy(CathRooms),copy.deepcopy(EPRooms),copy.deepcopy(procsNotPlaced)] for i in xrange(days)]
        self.weekBins = [[self.dayBins[i],self.dayBins[i+1],self.dayBins[i+2],self.dayBins[i+3],self.dayBins[i+4]] for i in xrange(0,len(self.dayBins)-4,5)]

        self.numCathRooms = numCathRooms
        self.numEPRooms = numEPRooms
        self.numDays = days
        self.numWeeks = len(self.weekBins)

        # statistical counters
        self.procsPlaced = 0
        self.crossOverProcs = 0
        self.cathToEP = 0           # procedures historically done in Cath that are scheduled in an EP room
        self.epToCath = 0           # procedures historically done in EP that are scheduled in a Cath room

    ##################################### BIN PACKING FOR #####################################
    #################################### WHOLE TIME PERIOD ####################################

    def packBins(self,procedures):
        '''
        Schedules procedures into the time period.
        
        Input: procedures (a list of cleaned procedure data for a given period of time)
        Returns: none
        '''

        allProcs = procedures[:]

        # break procedures up by scheduling horizon
        emergencies = [x for x in allProcs if x[iSchedHorizon]==1.0]
        sameDay = [x for x in allProcs if x[iSchedHorizon]==2.0]
        sameWeek = [x for x in allProcs if x[iSchedHorizon]==3.0]

        print "Total procedures after break up: "+str(len(emergencies)+len(sameDay)+len(sameWeek))
        print "Same day/emergencies: "+str(len(emergencies)+len(sameDay))
        print "Same weeks: "+str(len(sameWeek))
        
        
        # schedule emergencies and same day procedures first, day by day
        for d in range(1,timePeriod.numDays+1):
            #print "Packing bins for day: "+str(d)
            
            daysEmergencies = [proc for proc in emergencies if proc[iDay]==d]
            #print "On day "+str(d)+" there were "+str(len(daysEmergencies))+" emergency procedures"
            self.packBinsForDay(d-1,daysEmergencies)

            daysSameDays = [proc for proc in sameDay if proc[iDay]==d]
            #print "On day "+str(d)+" there were "+str(len(daysSameDays))+" same day procedures"
            self.packBinsForDay(d-1,daysSameDays)

        # schedule same week procedures next, week by week
        for w in range(1,timePeriod.numWeeks+1):

            weeksProcs = [proc for proc in sameWeek if proc[iWeek]==w]
            self.packBinsForWeek(w-1,weeksProcs)

        print "Total of "+str(self.procsPlaced)+" procedures placed"
            

    def sumProcTimes(self,dataList):
        '''
        Input: dataList (list of procedure data)
        
        Returns: the sum of the procedure times in dataList
        '''
        timeDataOnly = [dataList[i][iProcTime] for i in range(len(dataList))]
        
        return sum(timeDataOnly)

    ##################################### DAY BY DAY PACKING #####################################
    ################################### EMERGENCIES/SAME DAYS ####################################

    def getTotalMinutesInLabForDay(self,day,lab):
        '''
        Calculate the total number of minutes scheduled in a given lab for the
        day so far.
        
        Inputs: day (integer day of the time period, indexed from 0)
                lab (string name of the lab being queried)
        Returns: number of minutes scheduled into the lab that day (all rooms)
        '''
        total = 0
        if lab == 'Cath':
            CathRooms = self.dayBins[day][0]
            for i in xrange(self.numCathRooms):
                total += self.sumProcTimes(CathRooms[i])
        else:
            EPRooms = self.dayBins[day][0]
            for i in xrange(self.numEPRooms):
                total += self.sumProcTimes(EPRooms[i])

        return total

    def tryPlaceProcInLabDay(self,procedure,lab,day,nextOpenRoom,openRooms):
        '''
        Tries to place a procedure in a given room, if there is time for it in the schedule.
        Input: procedure (list of one procedure's data to be placed)
                lab (string name of lab to be scheduled into)
                day (integer day of time period, indexed from 0)
        Returns: True if placed, False otherwise
        '''

        if lab == 'Cath':
            rooms = self.dayBins[day][0]
        else:
            rooms = self.dayBins[day][1]
        
        potentialRoom = rooms[nextOpenRoom]
        potentialRoom.append(procedure)
        totalRoomTime = self.sumProcTimes(potentialRoom)

        if totalRoomTime > totalTimeRoom:
            potentialRoom.pop()
            openRooms.remove(nextOpenRoom)
            return False
        else:
            self.procsPlaced += 1
            return True

    def packBinsForDay(self,day,daysProcedures):
        '''
        Schedules procedures during a given day, if possible. Keeps track of overflow procedures
        (that couldn't be scheduled in that day).
        
        Input: day (integer day of time period to be scheduled, indexed from 0)
                daysProcedures (a list of procedure data for a given day)
        Returns: none
        '''
        procsNotPlaced = self.dayBins[day][2]

        #print "Packing same day procedures for day: "+str(day)
        #print "Number of procedures to pack: "+str(len(daysProcedures))

        # divides all procedures into inflexible ones (EP only, Cath only) and flexible ones (either lab)
        allProcs = daysProcedures[:]
        inflexibleProcs = [x for x in allProcs if x[iRoom]!=2.0]
        flexibleProcs = [y for y in allProcs if y[iRoom]==2.0]

        ##### PLACE ALL INFLEXIBLE PROCEDURES FIRST #####
        for i in xrange(len(inflexibleProcs)):

            openCathRooms = range(self.numCathRooms)
            openEPRooms = range(self.numEPRooms)

            procedure = inflexibleProcs[i]
            procRoom = roomConstraint[procedure[iRoom]]
            procPlaced = False

            while not procPlaced:

                # get the next rooms to be scheduled in both labs, if possible
                nextOpenCath = openCathRooms[0] if len(openCathRooms)>0 else -1
                nextOpenEP = openEPRooms[0] if len(openEPRooms)>0 else -1

                # procedure can be placed in CATH ONLY
                if procRoom == 'Cath':
                    # no rooms open in Cath: over capacity day
                    if nextOpenCath == -1:
                        procsNotPlaced.append(procedure)
                        procPlaced = True
                    else:
                        procPlaced = self.tryPlaceProcInLabDay(procedure,'Cath',day,nextOpenCath,openCathRooms)

                # procedure can be placed in EP ONLY
                elif procRoom == 'EP':
                    # no rooms open in EP: over capacity day
                    if nextOpenEP == -1:
                        procsNotPlaced.append(procedure)
                        procPlaced = True
                    else:
                        procPlaced = self.tryPlaceProcInLabDay(procedure,'EP',day,nextOpenEP,openEPRooms)


        
        ##### PLACE ALL FLEXIBLE PROCEDURES LAST #####
        for j in xrange(len(flexibleProcs)):

            openCathRooms = range(numCathRooms)
            openEPRooms = range(numEPRooms)
        
            procedure = flexibleProcs[j]
            originalLab = roomConstraint[procedure[iLab]]
            procPlaced = False

            while not procPlaced:

                # get the next rooms to be scheduled in both labs, if possible
                nextOpenCath = openCathRooms[0] if len(openCathRooms)>0 else -1
                nextOpenEP = openEPRooms[0] if len(openEPRooms)>0 else -1

                # no openings in either lab
                if (nextOpenEP == -1) and (nextOpenCath == -1):
                    procsNotPlaced.append(procedure)
                    procPlaced = True
                    
                # no openings in EP
                elif (nextOpenEP == -1):
                    procPlaced = self.tryPlaceProcInLabDay(procedure,'Cath',day,nextOpenCath,openCathRooms)
                    if procPlaced and originalLab=='EP':
                        self.crossOverProcs += 1
                        self.epToCath += 1

                # no openings in Cath
                elif (nextOpenCath == -1):
                    procPlaced = self.tryPlaceProcInLabDay(procedure,'EP',day,nextOpenEP,openEPRooms)
                    if procPlaced and originalLab=='Cath':
                        self.crossOverProcs += 1
                        self.cathToEP += 1
                        
                # openings in either lab
                else:
                    if (self.getTotalMinutesInLabForDay(day,'Cath')/(totalTimeRoom*self.numCathRooms)) <= (self.getTotalMinutesInLabForDay(day,'EP')/(totalTimeRoom*self.numEPRooms)):
                        procPlaced = self.tryPlaceProcInLabDay(procedure,'Cath',day,nextOpenCath,openCathRooms)
                        if procPlaced and originalLab=='EP':
                            self.crossOverProcs += 1
                            self.epToCath += 1
                    else:
                        procPlaced = self.tryPlaceProcInLabDay(procedure,'EP',day,nextOpenEP,openEPRooms)
                        if procPlaced and originalLab=='Cath':
                            self.crossOverProcs += 1
                            self.cathToEP += 1


                        

    ##################################### WEEK BY WEEK PACKING #####################################
    ################################### SAME WEEK PROCEDURES ONLY ##################################

    def getTotalMinutesInLabForWeek(self,week,lab):
        '''
        Calculate the total number of minutes scheduled in a given lab for the
        day so far.
        
        Inputs: day (integer day of the time period, indexed from 0)
                lab (string name of the lab being queried)
        Returns: number of minutes scheduled into the lab that week (all rooms)
        '''
        total = 0
        if lab == 'Cath':
            week = self.weekBins[week]
            for d in range(5):
                CathRooms = week[d][0]
                for i in xrange(self.numCathRooms):
                    total += self.sumProcTimes(CathRooms[i])
        else:
            week = self.weekBins[week]
            for d in range(5):
                EPRooms = week[d][1]
                for i in xrange(self.numEPRooms):
                    total += self.sumProcTimes(EPRooms[i])
        return total

    def tryPlaceProcInLabWeek(self,procedure,lab,week,nextOpenRoom,openRooms):
        '''
        Tries to place a procedure in a given room, if there is time for it in the week's schedule.
        Input: procedure (list of one procedure's data to be placed)
                lab (string name of lab to be scheduled into)
                week (integer week of time period, indexed from 0)
        Returns: True if placed, False otherwise
        '''
        weekRooms = self.weekBins[week]
        dayOfWeek = (nextOpenRoom/5)            # value:0-4 (0=M,1=T,2=W,3=R,4=F)
        nextOpenRoomOverall = nextOpenRoom      # with reference to the whole week
        
        if lab == 'Cath':
            nextOpenRoom = nextOpenRoom%self.numCathRooms       # with reference to just one day
            rooms = weekRooms[dayOfWeek][0]
        else:
            nextOpenRoom = nextOpenRoom%self.numEPRooms         # with reference to just one day
            rooms = weekRooms[dayOfWeek][1]
    
        potentialRoom = rooms[nextOpenRoom]
        potentialRoom.append(procedure)
        totalRoomTime = self.sumProcTimes(potentialRoom)

        if totalRoomTime > totalTimeRoom:
            potentialRoom.pop()
            openRooms.remove(nextOpenRoomOverall)
            return False
        else:
            self.procsPlaced += 1
            return True


    def packBinsForWeek(self,week,weeksProcedures):
        '''
        Schedules procedures during a given day, if possible. Keeps track of overflow procedures
        (that couldn't be scheduled in that week).
        
        Input: week (integer week of time period to be scheduled, indexed from 0)
                weeksProcedures (a list of procedure data for a given week)
        Returns: none
        '''
        
        weeksRooms = self.weekBins[week]
        # procedures that couldn't be placed in the entire week will be shown in the first day's (Monday's) overflow list
        procsNotPlaced = weeksRooms[0][2]

        # divides all procedures into inflexible ones (EP only, Cath only) and flexible ones (either lab)
        allProcs = weeksProcedures[:]
        inflexibleProcs = [x for x in allProcs if x[iRoom]!=2.0]
        flexibleProcs = [y for y in allProcs if y[iRoom]==2.0]

        ##### PLACE ALL INFLEXIBLE PROCEDURES FIRST #####
        for i in xrange(len(inflexibleProcs)):
            
            # all open rooms for the entire week (5 days)
            openCathRooms = range(self.numCathRooms*5) 
            openEPRooms = range(self.numEPRooms*5)

            procedure = inflexibleProcs[i]
            procRoom = roomConstraint[procedure[iRoom]]
            procPlaced = False

            while not procPlaced:

                # get the next rooms to be scheduled in both labs, if possible
                nextOpenCath = openCathRooms[0] if len(openCathRooms)>0 else -1
                nextOpenEP = openEPRooms[0] if len(openEPRooms)>0 else -1

                # procedure can be placed in CATH ONLY
                if procRoom == 'Cath':
                    # no rooms open in Cath: over capacity day
                    if nextOpenCath == -1:
                        print "overflow"
                        procsNotPlaced.append(procedure)
                        procPlaced = True
                    else:
                        procPlaced = self.tryPlaceProcInLabWeek(procedure,'Cath',week,nextOpenCath,openCathRooms)

                # procedure can be placed in EP ONLY
                elif procRoom == 'EP':
                    # no rooms open in EP: over capacity day
                    if nextOpenEP == -1:
                        procsNotPlaced.append(procedure)
                        procPlaced = True
                    else:
                        procPlaced = self.tryPlaceProcInLabWeek(procedure,'EP',week,nextOpenEP,openEPRooms)


        
        ##### PLACE ALL FLEXIBLE PROCEDURES LAST #####
        for j in xrange(len(flexibleProcs)):

            openCathRooms = range(numCathRooms*5)
            openEPRooms = range(numEPRooms*5)
        
            procedure = flexibleProcs[j]
            originalLab = roomConstraint[procedure[iLab]]
            procPlaced = False

            while not procPlaced:

                # get the next rooms to be scheduled in both labs, if possible
                nextOpenCath = openCathRooms[0] if len(openCathRooms)>0 else -1
                nextOpenEP = openEPRooms[0] if len(openEPRooms)>0 else -1

                # no openings in either lab
                if (nextOpenEP == -1) and (nextOpenCath == -1):
                    procsNotPlaced.append(procedure)
                    procPlaced = True
                    
                # no openings in EP
                elif (nextOpenEP == -1):
                    procPlaced = self.tryPlaceProcInLabWeek(procedure,'Cath',week,nextOpenCath,openCathRooms)
                    if procPlaced and originalLab=='EP':
                        self.crossOverProcs += 1
                        self.epToCath += 1

                # no openings in Cath
                elif (nextOpenCath == -1):
                    procPlaced = self.tryPlaceProcInLabWeek(procedure,'EP',week,nextOpenEP,openEPRooms)
                    if procPlaced and originalLab=='Cath':
                        self.crossOverProcs += 1
                        self.cathToEP += 1
                        
                # openings in either lab
                else:
                    if (self.getTotalMinutesInLabForWeek(week,'Cath')/(totalTimeRoom*self.numCathRooms*5)) <= (self.getTotalMinutesInLabForWeek(week,'EP')/(totalTimeRoom*self.numEPRooms*5)):
                        procPlaced = self.tryPlaceProcInLabWeek(procedure,'Cath',week,nextOpenCath,openCathRooms)
                        if procPlaced and originalLab=='EP':
                            self.crossOverProcs += 1
                            self.epToCath += 1
                    else:
                        procPlaced = self.tryPlaceProcInLabWeek(procedure,'EP',week,nextOpenEP,openEPRooms)
                        if procPlaced and originalLab=='Cath':
                            self.crossOverProcs += 1
                            self.cathToEP += 1

                                                           




######################################################################################################
######################################################################################################
##################################### READING/PROCESSING METHODS #####################################
######################################################################################################
######################################################################################################    

def readData(fileName):
    '''
    Input: fileName (string name of the file you want to process procedural data from

    Returns: a list of lists, each one being one procedure's information stored as floats
    '''
    procedures = []
    with open(fileName, 'rU') as f:
        reader = csv.reader(f)
        for row in reader:
            row = [float(i) for i in row[:numEntries+1]]
            procedures.append(row)
    
    return procedures


def cleanProcTimes(allProcs):
    '''
    Input: allProcs (list of all procedures as processed from csv)
    
    Returns: list of all procedures modified so that no procedure is
                of length zero, and procedures of length greater than
                totalTimeCath are truncated
    '''
    newProcs = allProcs[:]

    for i in range (len(newProcs)):
        procTime = newProcs[i][iProcTime]
        procTime += turnover
        if procTime > totalTimeRoom:
          procTime = totalTimeRoom
        newProcs[i][iProcTime] = procTime
    return newProcs


def getOptimizedTimeOnly(optimized):
    '''
    Based on a list of optimized scheduling, filter out irrelevant information and only
    include the procedure times.
    Input: optimized (a list, with each element corresponding to 1 day of scheduled procedures.
                Each element itself is a list, consisting of a dictionary with Cath rooms
                and their assigned procedures, a dictionary with EP rooms and their
                assigned procedures, and a list of procedures that went over capacity.
                This should be the output of packBins(procedures)
    Returns: a copy of the same input list, but only including procedure times
    '''

    optimizedCopy = optimized[:]

    for d in xrange(len(optimized)):
        Cath = optimized[d][0]
        EP = optimized[d][1]
        overCapacity = optimized[d][2]
        for c in xrange(numCathRooms):
            day = Cath[c]
            timesOnly = [x[iProcTime] for x in day]
            Cath[c] = timesOnly
        for e in xrange(numEPRooms):
            day = EP[e]
            timesOnly = [x[iProcTime] for x in day]
            EP[e] = timesOnly
        optimized[d][2] = [x[iProcTime] for x in overCapacity]
    return optimizedCopy


def cleanResults(optimized):
    '''
    Input: optimizedTimeOnly (a list of the optimized procedure times. Should be output from
            getOptimizedTimeOnly(optimized)
    Returns: a copy of the same input list, but with all days equalized in terms of number of
                procedures listed per room day and overflow procedures per day
    '''
    optimizedCopy = optimized[:]
    
    # get the OVERALL maximum number of procedures in Cath/EP/Overflow for the time period
    maxNumProcsCathRoom = 0
    maxNumProcsEPRoom = 0
    maxNumOverflow = 0
    for d in xrange(len(optimized)):

        CathDict = optimizedCopy[d][0]
        EPDict = optimizedCopy[d][1]
        overflowList = optimizedCopy[d][2]

        for c in xrange(numCathRooms):
            room = CathDict[c]
            maxNumProcsCathRoom = len(room) if len(room)>maxNumProcsCathRoom else maxNumProcsCathRoom
        for e in xrange(numEPRooms):
            room = EPDict[e]
            maxNumProcsEPRoom = len(room) if len(room)>maxNumProcsEPRoom else maxNumProcsEPRoom
            
        maxNumOverflow = len(overflowList) if len(overflowList)>maxNumOverflow else maxNumOverflow
 
    # add elements to each room day if necessary, so each room day has the same number of elements
    for d in xrange(len(optimized)):

        CathDict = optimizedCopy[d][0]
        EPDict = optimizedCopy[d][1]
        
        for c in xrange(numCathRooms):
            room = CathDict[c]
            numDiffProcs = maxNumProcsCathRoom - len(room)
            room += numDiffProcs*[0.00]

        for e in xrange(numEPRooms):
            room = EPDict[e]
            numDiffProcs = maxNumProcsEPRoom - len(room)
            room += numDiffProcs*[0.00]
        
        overflowList = optimizedCopy[d][2]
        numDiffProcs = maxNumOverflow - len(overflowList)
        overflowList += numDiffProcs*[0.00]

    return optimizedCopy

        

def saveResults(cleanOptimizedTimeOnly, workbook, sheet):
    '''
    Saves the results of the optimization to an Excel sheet organized by day (rows)
    and procedure room (Cath Room # Proc #, EP Room # Proc #, Overflow Proc #;  columns)
    '''
    
    wb = px.load_workbook(workbook)
    entrySheet = wb.get_sheet_by_name(name = sheet)

    maxNumCathProcs = len(cleanOptimizedTimeOnly[0][0][0])
    maxNumEPProcs = len(cleanOptimizedTimeOnly[0][1][0])

    numCathColumns = numCathRooms*maxNumCathProcs
    numEPColumns = numEPRooms*maxNumEPProcs
    numOverflowColumns = len(cleanOptimizedTimeOnly[0][2])

    # write column labels
    entrySheet.cell(row=0,column=0).value = "Day of Period"
    for cath in xrange(numCathRooms):
        for proc in xrange(1,maxNumCathProcs+1):
            entrySheet.cell(row=0,column=(maxNumCathProcs*cath)+proc).value = "Cath Room "+str(cath+1)+" Proc "+str(proc)
    for ep in xrange(numEPRooms):
        for proc in xrange(1,maxNumEPProcs+1):
            entrySheet.cell(row=0,column=(numCathColumns+((maxNumEPProcs*ep)+proc))).value = "EP Room "+str(ep+1)+" Proc "+str(proc)
    for o in xrange(1,numOverflowColumns+1):
        entrySheet.cell(row=0, column=numCathColumns+numEPColumns+o).value = "Overflow Proc "+str(o)


    # write days' information
    for d in xrange(1,len(cleanOptimizedTimeOnly)+1):

        # write day label
        entrySheet.cell(row=d,column=0).value = str(d)

        # write room-procedure/overflow-procedure information
        for cath in xrange(numCathRooms):
            for proc in xrange(1,maxNumCathProcs+1):
                entrySheet.cell(row=d,column=(maxNumCathProcs*cath)+proc).value = str(round(cleanOptimizedTimeOnly[d-1][0][cath][proc-1],2))
        for ep in xrange(numEPRooms):
            for proc in xrange(1,maxNumEPProcs+1):
                entrySheet.cell(row=d,column=(numCathColumns+((maxNumEPProcs*ep)+proc))).value = str(round(cleanOptimizedTimeOnly[d-1][1][ep][proc-1],2))
        for o in xrange(1,numOverflowColumns+1):
            entrySheet.cell(row=d, column=numCathColumns+numEPColumns+o).value = str(round(cleanOptimizedTimeOnly[d-1][2][o-1],2))

    wb.save(workbook)

    print "Finished saving results: please see "+workbook+ " for results."
        

if __name__ == "__main__":


    ############# VERIFY FOLLOWING VALUES BEFORE RUNNING ##############

    # information regarding the scheduling parameters:
    totalTimeRoom = 10.58*60    # total time available in a room per day (min)
    closeCap = 10*60            # time cap for closing a room (min)
    turnover = 0                # estimated time for room turnover (min)
    numCathRooms = 5            # number of Cath rooms available per day
    numEPRooms = 3              # number of EP rooms available per day

    ###### information regarding the order of information in the data sheet ######
    numEntries = 6              # number of columns in data sheet
    iDay = 0                    # index: Day of period
    iWeek = 1                   # index: Week of period
    iLab = 2                    # index: Lab key (Cath,EP)
    iProcTime = 3               # index: Procedure time (minutes)
    iRoom = 4                   # index: Room constraint (Cath only, EP only, either)
    iProcType = 6               # index: Procedure type key
    iSchedHorizon = 5           # index: Scheduling horizon key
    daysInPeriod = 125          # integer: Number of days in period
    
    roomConstraint = {0.0:'Cath', 1.0:'EP', 2.0:'either'}
                                # room constraint key used in data file

    ###### information regarding the name/location of the data file ######
    
    os.chdir("/Users/nicseo/Desktop/MIT/Junior/Fall/UROP/Scheduling Optimization/Script")
    # uncomment the data set to analyze or add a new one
    #fileName= 'CathAndEP_PT_SchedHorizon.csv'
    fileName= 'Volumes1.csv'
    #fileName= 'Volumes2.csv'

    ###### information regarding the name/location of the output data ######
    ###### which must be created before running this script ######
    
    # uncomment the workbook to save to or add a new one
    workbook = "OptimizedSchedule_RoomHorizonConstraints.xlsx"
    #workbook = "OptimizedSchedule.xlsx"
    #workbook = "Optimized_VolumeIncrease.xlsx"
    sheet = "Results"
    

    ############# RUNNING OF THE SCRIPT: not necessary to modify #############

    ###### read/process in data ######
    procedures = readData(fileName)
    procedures = cleanProcTimes(procedures)
    
    ###### model time period / pack bins ######
    timePeriod = TimePeriod(daysInPeriod,numCathRooms,numEPRooms)   # time period model
    timePeriod.packBins(procedures)                                 # schedule all procedures from data file

    overflowWeeks = 0
    for week in timePeriod.weekBins:
        day = week[0]
        if len(day[2]) > 0:
            overflowWeeks += 1
    print "Total number of overflow weeks: "+str(overflowWeeks)

    overflowProcs = 0
    for day in timePeriod.dayBins:
        overflowProcs += len(day[2])
    print "Total number of overflow procs: "+str(overflowProcs)

    print "Total number of crossover procedures: "+str(timePeriod.crossOverProcs)
    print "Total number of Cath procedures in EP: "+str(timePeriod.cathToEP)
    print "Total number of EP procedures in Cath: "+str(timePeriod.epToCath)
    
    # process results
    optimized = copy.deepcopy(timePeriod.dayBins)
    optimizedTimeOnly = getOptimizedTimeOnly(optimized)             # only look at proc times
    cleanedOptimized = cleanResults(optimizedTimeOnly)              # format results for excel

    #saveResults(cleanedOptimized,workbook,sheet)
    # comment out previous line if you just want to look at summary stats in the shell,
    # and not change the excel sheet

