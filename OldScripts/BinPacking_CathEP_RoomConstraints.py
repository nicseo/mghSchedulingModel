"""
Created on Thursday, September 18, 2014

@author: nicseo

Notes about use:
- The first part of this script sets up the necessary functions to carry out the analysis.
    At the bottom of this file, in the __main__ function, there are defined some variables
    that the user should verify before running the script (e.g. regarding the csv that the
    data is being read from).

- To read the final output, index the returned list by DAY, then by LAB KEY,
    then by ROOM NUMBER (optional).
    e.g. results[3][0] gives Cath's optimized schedule on the 4th day of the period.

- Results are also output to an Excel sheet, whose workbook name and sheet must EXIST and
    be specified before running the script. Make sure the sheet is in the same directory
    as this script.
    
- If you change the number of Cath/EP rooms in the scheduling parameters, it is recommended
    that you CLEAR the output file of contents AND formatting before running.

"""
import csv
import os
import openpyxl as px

def readData(fileName):
    '''
    Input: fileName (string name of the file you want to process procedural data from

    Returns: a list of lists, each one being one procedure's information stored as floats
    '''
    procedures = []
    with open(fileName, 'rU') as f:
        reader = csv.reader(f)
        for row in reader:
            row = [float(i) for i in row[:6]]
            procedures.append(row)
    
    return procedures


def cleanProcTimes(allProcs):
    '''
    Input: allProcs (list of all procedures as processed from csv)
    
    Returns: list of all procedures modified so that no procedure is
                of length zero, and procedures of length greater than
                totalTimeCath are truncated
    '''
    newProcs = allProcs[:]

    for i in range (len(newProcs)):
        procTime = newProcs[i][iProcTime]
        procTime += turnover
        if procTime > totalTimeRoom:
          procTime = totalTimeRoom
        newProcs[i][iProcTime] = procTime
    return newProcs

def sumProcTimes(dataList):
    '''
    Input: dataList (list of procedure data as processed from csv)
    
    Returns: the sum of the procedure times in dataList
    '''
    timeDataOnly = [dataList[i][iProcTime] for i in range(len(dataList))]
    
    return sum(timeDataOnly)


def tryPlaceProcInLab(procedure,rooms,nextOpenRoom,openRooms,procsPlaced):

    '''
    Tries to place a procedure in a given room, if there is time for it in the schedule.
    Input: procedure (list of one procedure's data to be placed)
            rooms (dictionary containing the lab rooms)
            nextOpenRoom (the next room to be scheduled into)
            closeCap (the limit for minutes to be scheduled in the room)
            procsPlaced (the number of minutes placed in the lab so far)
    Returns: True if placed, False otherwise, and the new number of minutes placed in the lab so far
    '''
    potentialRoom = rooms[nextOpenRoom]
    potentialRoom.append(procedure)
    totalRoomTime = sumProcTimes(potentialRoom)

    if totalRoomTime > totalTimeRoom:
        potentialRoom.pop()
        openRooms.remove(nextOpenRoom)
        return (False,procsPlaced)
    else:
        procsPlaced += procedure[iProcTime]
        return (True,procsPlaced)

def packBinsForDay(daysProcedures):
    '''
    Input: daysProcedures (a list of procedure data for a given day)

    Returns: a list whose first item is a dictionary with Cath rooms
                and their assigned procedures, whose second item
                is a dictionary with EP rooms and their assigned
                procedures, and whose third item is a list of procedures
                that could not fit into that day's schedule
    '''
    CathRooms = {i:[] for i in xrange(numCathRooms)}
    EPRooms = {i:[] for i in xrange(numEPRooms)}
    # keeps track of procedures that could not fit into a day: over capacity
    procsNotPlaced = []
    
    # keeps track of the procedure time placed in each lab so far that day
    cathProcsPlaced = 0
    epProcsPlaced = 0

    # divides all procedures into inflexible ones (EP only, Cath only) and flexible ones (either lab)
    allDaysProcs = daysProcedures[:]
    inflexibleProcs = [x for x in allDaysProcs if x[iRoom]!=2.0]
    flexibleProcs = [y for y in allDaysProcs if y[iRoom]==2.0]

    ##### PLACE ALL INFLEXIBLE PROCEDURES FIRST #####
    for i in xrange(len(inflexibleProcs)):

        openCathRooms = range(numCathRooms)
        openEPRooms = range(numEPRooms)

        procedure = inflexibleProcs[i]
        procTime = procedure[iProcTime]
        procRoom = roomConstraint[procedure[iRoom]]
        procPlaced = False

        while not procPlaced:

            # get the next rooms to be scheduled in both labs, if possible
            nextOpenCath = openCathRooms[0] if len(openCathRooms)>0 else -1
            nextOpenEP = openEPRooms[0] if len(openEPRooms)>0 else -1

            # procedure can be placed in CATH ONLY
            if procRoom == 'Cath':
                # no rooms open in Cath: over capacity day
                if nextOpenCath == -1:
                    procsNotPlaced.append(procedure)
                    procPlaced = True
                else:
                    procPlaced,cathProcsPlaced = tryPlaceProcInLab(procedure,CathRooms,nextOpenCath,openCathRooms,cathProcsPlaced)

            # procedure can be placed in EP ONLY
            elif procRoom == 'EP':
                # no rooms open in EP: over capacity day
                if nextOpenEP == -1:
                    procsNotPlaced.append(procedure)
                    procPlaced = True
                else:
                    procPlaced,epProcsPlaced = tryPlaceProcInLab(procedure,EPRooms,nextOpenEP,openEPRooms,epProcsPlaced)


    
    ##### PLACE ALL FLEXIBLE PROCEDURES LAST #####
    for j in xrange(len(flexibleProcs)):

        openCathRooms = range(numCathRooms)
        openEPRooms = range(numEPRooms)
    
        procedure = flexibleProcs[j]
        procTime = procedure[iProcTime]
        procPlaced = False

        while not procPlaced:

            # get the next rooms to be scheduled in both labs, if possible
            nextOpenCath = openCathRooms[0] if len(openCathRooms)>0 else -1
            nextOpenEP = openEPRooms[0] if len(openEPRooms)>0 else -1

            # no openings in either lab
            if (nextOpenEP == -1) and (nextOpenCath == -1):
                procsNotPlaced.append(procedure)
                procPlaced = True
                
            # no openings in EP
            elif (nextOpenEP == -1):
                procPlaced,cathProcsPlaced = tryPlaceProcInLab(procedure,CathRooms,nextOpenCath,openCathRooms,cathProcsPlaced)

            # no openings in Cath
            elif (nextOpenCath == -1):
                procPlaced,epProcsPlaced = tryPlaceProcInLab(procedure,EPRooms,nextOpenEP,openEPRooms,epProcsPlaced)
                    
            # openings in either lab
            else:
                if (cathProcsPlaced/(totalTimeRoom*numCathRooms)) <= (epProcsPlaced/(totalTimeRoom*numEPRooms)):
                    procPlaced,cathProcsPlaced = tryPlaceProcInLab(procedure,CathRooms,nextOpenCath,openCathRooms,cathProcsPlaced)
                else:
                    procPlaced,epProcsPlaced = tryPlaceProcInLab(procedure,EPRooms,nextOpenEP,openEPRooms,epProcsPlaced)

    return [CathRooms, EPRooms, procsNotPlaced]

def packBins(procedures):
    '''
    Input: procedures (a list of cleaned procedure data for a given period of time)

    Returns: a list, with each element corresponding to 1 day of scheduled procedures.
                Each element itself is a list, consisting of a dictionary with Cath rooms
                and their assigned procedures, a dictionary with EP rooms and their
                assigned procedures, and a list of procedures that went over capacity

    NOTE: to read output, index by DAY, then by LAB, then by ROOM NUMBER
    '''

    allProcs = procedures[:]
    optimizedDays = []

    for d in range(1,daysInPeriod+1):
        daysProcedures = [proc for proc in allProcs if proc[iDay]==d]
        optimizedDays.append(packBinsForDay(daysProcedures))

    return optimizedDays


def getOptimizedTimeOnly(optimized):
    '''
    Based on a list of optimized scheduling, filter out irrelevant information and only
    include the procedure times.
    Input: optimized (a list, with each element corresponding to 1 day of scheduled procedures.
                Each element itself is a list, consisting of a dictionary with Cath rooms
                and their assigned procedures, a dictionary with EP rooms and their
                assigned procedures, and a list of procedures that went over capacity.
                This should be the output of packBins(procedures)
    Returns: a copy of the same input list, but only including procedure times
    '''

    optimizedCopy = optimized[:]

    for d in xrange(len(optimized)):
        Cath = optimized[d][0]
        EP = optimized[d][1]
        overCapacity = optimized[d][2]
        for c in xrange(numCathRooms):
            day = Cath[c]
            timesOnly = [x[iProcTime] for x in day]
            Cath[c] = timesOnly
        for e in xrange(numEPRooms):
            day = EP[e]
            timesOnly = [x[iProcTime] for x in day]
            EP[e] = timesOnly
        optimized[d][2] = [x[iProcTime] for x in overCapacity]
    return optimizedCopy


def cleanResults(optimizedTimeOnly):
    '''
    Input: optimizedTimeOnly (a list of the optimized procedure times. Should be output from
            getOptimizedTimeOnly(optimized)
    Returns: a copy of the same input list, but with all days equalized in terms of number of
                procedures listed per room day and overflow procedures per day
    '''
    optimizedCopy = optimized[:]
    
    # get the OVERALL maximum number of procedures in Cath/EP/Overflow for the time period
    maxNumProcsCathRoom = 0
    maxNumProcsEPRoom = 0
    maxNumOverflow = 0
    for d in xrange(len(optimized)):

        CathDict = optimizedCopy[d][0]
        EPDict = optimizedCopy[d][1]
        overflowList = optimizedCopy[d][2]

        for c in xrange(numCathRooms):
            room = CathDict[c]
            maxNumProcsCathRoom = len(room) if len(room)>maxNumProcsCathRoom else maxNumProcsCathRoom
        for e in xrange(numEPRooms):
            room = EPDict[e]
            maxNumProcsEPRoom = len(room) if len(room)>maxNumProcsEPRoom else maxNumProcsEPRoom
            
        maxNumOverflow = len(overflowList) if len(overflowList)>maxNumOverflow else maxNumOverflow
 
    # add elements to each room day if necessary, so each room day has the same number of elements
    for d in xrange(len(optimized)):

        CathDict = optimizedCopy[d][0]
        EPDict = optimizedCopy[d][1]
        
        for c in xrange(numCathRooms):
            room = CathDict[c]
            numDiffProcs = maxNumProcsCathRoom - len(room)
            room += numDiffProcs*[0.00]

        for e in xrange(numEPRooms):
            room = EPDict[e]
            numDiffProcs = maxNumProcsEPRoom - len(room)
            room += numDiffProcs*[0.00]
        
        overflowList = optimizedCopy[d][2]
        numDiffProcs = maxNumOverflow - len(overflowList)
        overflowList += numDiffProcs*[0.00]

    return optimizedCopy

        

def saveResults(cleanOptimizedTimeOnly, workbook, sheet):
    '''
    Saves the results of the optimization to an Excel sheet organized by day (rows)
    and procedure room (Cath Room # Proc #, EP Room # Proc #, Overflow Proc #;  columns)
    '''
    
    wb = px.load_workbook(workbook)
    entrySheet = wb.get_sheet_by_name(name = sheet)

    maxNumCathProcs = len(cleanOptimizedTimeOnly[0][0][0])
    maxNumEPProcs = len(cleanOptimizedTimeOnly[0][1][0])

    numCathColumns = numCathRooms*maxNumCathProcs
    numEPColumns = numEPRooms*maxNumEPProcs
    numOverflowColumns = len(cleanOptimizedTimeOnly[0][2])

    # write column labels
    entrySheet.cell(row=0,column=0).value = "Day of Period"
    for cath in xrange(numCathRooms):
        for proc in xrange(1,maxNumCathProcs+1):
            entrySheet.cell(row=0,column=(maxNumCathProcs*cath)+proc).value = "Cath Room "+str(cath+1)+" Proc "+str(proc)
    for ep in xrange(numEPRooms):
        for proc in xrange(1,maxNumEPProcs+1):
            entrySheet.cell(row=0,column=(numCathColumns+((maxNumEPProcs*ep)+proc))).value = "EP Room "+str(ep+1)+" Proc "+str(proc)
    for o in xrange(1,numOverflowColumns+1):
        entrySheet.cell(row=0, column=numCathColumns+numEPColumns+o).value = "Overflow Proc "+str(o)


    # write days' information
    for d in xrange(1,len(cleanOptimizedTimeOnly)+1):

        # write day label
        entrySheet.cell(row=d,column=0).value = str(d)

        # write room-procedure/overflow-procedure information
        for cath in xrange(numCathRooms):
            for proc in xrange(1,maxNumCathProcs+1):
                entrySheet.cell(row=d,column=(maxNumCathProcs*cath)+proc).value = str(round(cleanOptimizedTimeOnly[d-1][0][cath][proc-1],2))
        for ep in xrange(numEPRooms):
            for proc in xrange(1,maxNumEPProcs+1):
                entrySheet.cell(row=d,column=(numCathColumns+((maxNumEPProcs*ep)+proc))).value = str(round(cleanOptimizedTimeOnly[d-1][1][ep][proc-1],2))
        for o in xrange(1,numOverflowColumns+1):
            entrySheet.cell(row=d, column=numCathColumns+numEPColumns+o).value = str(round(cleanOptimizedTimeOnly[d-1][2][o-1],2))

    wb.save(workbook)

    print "Finished saving results: please see "+workbook+ " for results."
        

if __name__ == "__main__":


    ############# VERIFY FOLLOWING VALUES BEFORE RUNNING ##############

    # information regarding the scheduling parameters:
    totalTimeRoom = 10.58*60    # total time available in a room per day (min)
    closeCap = 10*60            # time cap for closing a room (min)
    turnover = .58*60           # estimated time for room turnover (min)
    numCathRooms = 5            # number of Cath rooms available per day
    numEPRooms = 4              # number of EP rooms available per day


    # information regarding the order of information in the data sheet:
    iDay = 0                    # index: Day of period
    iWeek = 1                   # index: Week of period
    iLab = 2                    # index: Lab key (Cath,EP)
    iProcTime = 3               # index: Procedure time (minutes)
    iRoom = 4                   # index: Room constraint (Cath only, EP only, either)
    iProcType = 5               # index: Procedure type key
    daysInPeriod = 125          # integer: Number of days in period
    roomConstraint = {0.0:'Cath', 1.0:'EP', 2.0:'either'}
                                # room constraint key used in data file


    # information regarding the name/location of the data file:
    os.chdir("/Users/nicseo/Desktop/MIT/Junior/Fall/UROP/Scheduling Optimization")
    fileName= 'CathAndEP_Primetime.csv'


    # information regarding the name/location of the output data, which must
    # be created before running this script
    workbook = "OptimizedSchedule.xlsx"
    sheet = "Results"

    

    ############# RUNNING OF THE SCRIPT: not necessary to modify #############
    
    
    procedures = readData(fileName)
    procedures = cleanProcTimes(procedures)
    optimized = packBins(procedures)
    optimizedTimeOnly = getOptimizedTimeOnly(optimized)
    cleanedOptimized = cleanResults(optimizedTimeOnly)
    saveResults(cleanedOptimized,workbook,sheet)

